###### CONTROL ######
config = "api/config.txt"
debug = open(config).read().replace("\n", "")[open(config).read().replace("\n", "").index("debug") - 5]
mobile = open(config).read().replace("\n", "")[open(config).read().replace("\n", "").index("mobile") - 5]
#-------------------#
from math import *
if debug == "Y":
    import apiHandler
    import asyncApi
else:
    from api import apiHandler
    from api import asyncApi
import os
os.environ["SDL_VIDEODRIVER"] = "dummy"
import pygame
import base64
from datetime import datetime
from PIL import Image
from openpyxl import Workbook
from io import BytesIO

###### OPERATOR FUNCTIONS ######
def getDays(date):
    monthLengths = [31,28,31,30,31,30,31,31,30,31,30,31]
    return int(date[:4]) * 365 + sum(monthLengths[:int(date[5:7]) - 1]) + int(date[8:])

def setLength(value, numDigits):
    zeroString = ""
    if len(str(value)) < numDigits:
        for i in range(numDigits - len(str(value))):
            zeroString = zeroString + "0"
    return zeroString + str(value)

def createPlot(data, teamname, startDate):
    pygame.init()
    windowSize = [630, 250]
    screen = pygame.display.set_mode(windowSize)

    ranks = pygame.image.load("ranks.png")
    point = pygame.image.load("point.png")

    font = pygame.font.Font("Teko-Regular.ttf", 22)
    font2 = pygame.font.Font("Teko-Regular.ttf", 18)

    #screen.fill((14, 97, 114))
    screen.fill((8, 36, 42))
    topMargin = 25
    bottomMargin = 50
    chartDimensions = [windowSize[0] - 50, windowSize[1] - (topMargin + bottomMargin)]
    vertTicks, horizTicks = [5, 6]

    leftBound = (windowSize[0] - chartDimensions[0]) / 2
    rightBound = (windowSize[0] + chartDimensions[0]) / 2
    topBound = topMargin
    bottomBound = windowSize[1] - bottomMargin

    start = getDays(startDate)
    now = f"{datetime.now().year}-{setLength(datetime.now().month, 2)}-{setLength(datetime.now().day, 2)}"
    today = getDays(now)
    longBreak = False

    if today - getDays(data[-1][0]) > 50:
        today = getDays(data[-1][0]) + 5
        now = data[-1][0]
        longBreak = True

    ranks = pygame.transform.scale(ranks, (chartDimensions[0], chartDimensions[1]))
    ranksRect = ((leftBound, topBound), (ranks.get_rect()[2], ranks.get_rect()[3]))
    screen.blit(ranks, ranksRect)
    
    for horizLine in range(horizTicks):
        pygame.draw.aaline(screen, (44, 127, 144), (leftBound, topBound + horizLine * (chartDimensions[1] / (horizTicks - 1))), (rightBound, topBound + horizLine * (chartDimensions[1] / (horizTicks - 1))))

    xValues = [getDays(item[0]) for item in data]
    yValues = [item[1] for item in data]
    nextXPos = (xValues[0] - start) / (today - start) * chartDimensions[0] + leftBound
    nextYPos = (1 - yValues[0] / 100) * (chartDimensions[1]) + topBound
    for index in range(len(data) - 1):
        xPos = (xValues[index] - start) / (today - start) * chartDimensions[0] + leftBound
        yPos = (1 - yValues[index] / 100) * (chartDimensions[1]) + topBound
        nextXPos = (xValues[min(index + 1, len(xValues) - 1)] - start) / (today - start) * chartDimensions[0] + leftBound
        nextYPos = (1 - yValues[min(index + 1, len(yValues) - 1)] / 100) * (chartDimensions[1]) + topBound

        #pygame.draw.aaline(screen, (127,160,160), (xPos, yPos), (nextXPos - 3, yPos), 6)
        pygame.draw.aaline(screen, (255,255,255), (xPos, yPos), (nextXPos - 3, yPos), 4)
        #pygame.draw.aaline(screen, (127,160,160), (nextXPos - 3, yPos), (nextXPos, nextYPos), 6)
        pygame.draw.aaline(screen, (255,255,255), (nextXPos - 3, yPos), (nextXPos, nextYPos),4)
    pygame.draw.aaline(screen, (255,255,255), (nextXPos, nextYPos), (rightBound, nextYPos),4)

    screen.blit(point, (rightBound - 8, nextYPos - 8))
    bottomText = font.render(f"{teamname} Powerscore Over Time", True, (255,255,255))
    bottomRect = [windowSize[0]/2 - bottomText.get_rect()[2]/2, windowSize[1] - 40, bottomText.get_rect()[2], bottomText.get_rect()[3]]
    screen.blit(bottomText, bottomRect)

    startDateText = font2.render(startDate, True, (255,255,255))
    startDateRect = [20, windowSize[1] - 40, startDateText.get_rect()[2], startDateText.get_rect()[3]]
    screen.blit(startDateText, startDateRect)

    if longBreak:
        endDateText = font2.render(f"Last Active ({now})", True, (255,255,255))
    else:
        endDateText = font2.render(f"Now ({now})", True, (255,255,255))
    endDateRect = [windowSize[0] - endDateText.get_rect()[2] - 20, windowSize[1] - 40, endDateText.get_rect()[2], endDateText.get_rect()[3]]
    screen.blit(endDateText, endDateRect)
    #pygame.draw.circle(screen, (255, 255, 255), (rightBound, yPos), 2)

    plotBytes = BytesIO()
    pygame.image.save(screen, plotBytes, "jpg")
    plotBytes.seek(0)
    return plotBytes


###### RANKING ######

ranking = [
    ["Bronze I", 0],
    ["Bronze II", 11],
    ["Bronze III", 21],

    ["Silver I", 31],
    ["Silver II", 38],
    ["Silver III", 44],

    ["Gold I", 51],
    ["Gold II", 58],
    ["Gold III", 64],

    ["Diamond I", 71],
    ["Diamond II", 74],
    ["Diamond III", 78],

    ["Royal I", 81],
    ["Royal II", 84],
    ["Royal III", 88],

    ["Ethereal I", 91],
    ["Ethereal II", 94],
    ["Ethereal III", 97],
    ["Upper Limit", 100]
]

def giveRanking(value): # function to get the rank given a powerscore
    index = 0
    while ranking[index][1] < value:
        index += 1
    index -= 1
    return ranking[index][0]

###### ALGORITHM ######

def runPowerScore(compName, compID, div, typeOfPowerscore, compInfo, onlyForComp=False, scalingFactor=1):
    #output = open("output.txt", "w")

    def generateTeamsListFromComp(fileList): # uses the match list to generate a list of teams that attended a comp, so that a second teamList is unnecessary
        uniqueTeams = []
        for row in fileList:
            teamsInRow = row[:4]
            for team in teamsInRow:
                if not team in uniqueTeams:
                    uniqueTeams.append(team)

        return uniqueTeams

    chart = apiHandler.getMatchList(compName, compInfo)
    teamList = generateTeamsListFromComp(chart)

    ###### FUNCTIONS ######
    def retrieve(round=None, player=None): # retrieves a player's matches from the dataset, or optionally the data of a specific match (although the powerscore algorithm never uses this 2nd option)
        if round != None:
            specifiedValue = "N/A"
            for row in chart:
                if row[0] == round:
                    specifiedValue = row
            return specifiedValue
        elif player != None:
            foundMatches = []
            for row in chart:
                if player in row:
                    foundMatches.append(row)
            return foundMatches

    def getListOfDiffs(player): # returns a list of all the score differentials in matches that a specific player had (score differential is the difference in score between alliance and opponent)
        matchList = retrieve(player=player)

        scoreDiffs = []
        for match in matchList:
            onRed = False
            if match.index(player) in [0,1]:
                onRed = True

            if match[4] == "":
                match[4] = 0
            if match[5] == "":
                match[5] = 0
                
            if onRed:
                scoreDiff = int(match[4]) - int(match[5])
            else:
                scoreDiff = int(match[5]) - int(match[4])
            scoreDiffs.append(scoreDiff * scalingFactor)
        
        return scoreDiffs
        
    def getAverageDiff(player): # returns the average value of a list of score differentials. this is separated from the above function because we sometimes want the list, other times want the avg
        scoreDiffs = getListOfDiffs(player=player)
        avgDiff = sum(scoreDiffs) / len(scoreDiffs)
        return round(avgDiff*1000)/1000

    def getAllianceAverageDiffs(player): # gets a list of the average score differentials amongst all the alliance partners a team has had. used to compare against the player's own score diffs
        matchList = retrieve(player=player)
        listOfAllianceAvgDiffs = []

        for match in matchList: # retrieves the index within the row of the alliance partner, given a team
            if match.index(player) in [0,1]:
                alliancePartner = match[(match.index(player) == 0) + 0]
            elif match.index(player) in [2,3]:
                alliancePartner = match[(match.index(player) == 2) + 2]
            listOfAllianceAvgDiffs.append(getAverageDiff(alliancePartner))

        return listOfAllianceAvgDiffs

    def getOpponentAverageDiffs(player): # gets a list of the average score differentials amongst all the alliance partners a team has had. used to compare against the player's own score diffs
        matchList = retrieve(player=player)
        listOfOpponentAvgDiffs = []

        for match in matchList: # retrieves the index within the row of the two opponent teams, given a team
            if match.index(player) in [0,1]:
                opponent1 = match[2]
                opponent2 = match[3]
            elif match.index(player) in [2,3]:
                opponent1 = match[0]
                opponent2 = match[1]
            avgDiff1 = getAverageDiff(opponent1)
            avgDiff2 = getAverageDiff(opponent2)
            
            listOfOpponentAvgDiffs.append((avgDiff1 + avgDiff2) / 2) # averages the opponent match diffs

        return listOfOpponentAvgDiffs

    def getPowerScore(player, typeOfPowerscore): # master function, combines all the above functions to generate a powerScore value that reflects how "powerful" a team was in the competition
        matchDiffs = getListOfDiffs(player=player) # factor 1 (positive)
        allianceDiffs = getAllianceAverageDiffs(player=player) # factor 2 (negative)
        opponentDiffs = getOpponentAverageDiffs(player=player) # factor 3 (positive)
        matchPowers = []

        for index, matchDiff in enumerate(matchDiffs):
            if typeOfPowerscore == "offensive":
                if onlyForComp:
                    matchPower = (2/(1 + exp(-0.05 * (matchDiff + opponentDiffs[index])))) - 1 # performs sigmoid calculation based on the 2 factors
                else:
                    matchPower = matchDiff + opponentDiffs[index] # performs summation of 2 factors
            if typeOfPowerscore == "defensive":
                if onlyForComp:
                    matchPower = (2/(1 + exp(-0.05 * (matchDiff - allianceDiffs[index])))) - 1 # performs sigmoid calculation based on the 2 factors
                else:
                    matchPower = matchDiff - allianceDiffs[index] # performs summation of 2 factors
            if typeOfPowerscore == "general":
                if onlyForComp:
                    matchPower = (2/(1 + exp(-0.05 * (matchDiff - allianceDiffs[index] + opponentDiffs[index])))) - 1 # performs sigmoid calculation based on the 3 factors
                else:
                    matchPower = matchDiff - allianceDiffs[index] + opponentDiffs[index] # performs summation of 3 factors
            matchPowers.append(matchPower)
        
        if onlyForComp:
            powerScore = (sum(matchPowers) / len(matchPowers)) / 2 + 0.5
            return round(powerScore * 1000) / 10 # change to 10 outside the parentheses if need to revert
        else:
            powerScore = (sum(matchPowers) / len(matchPowers))# / 2 + 0.5
            return round(powerScore * 1000) / 1000 # change to 10 outside the parentheses if need to revert
            
    ###### MAIN ######
    #print(teamList)
    for team in teamList: # changes each item in the team list to be a tuple of the team name and its powerScore as opposed to just the team name
        teamList[teamList.index(team)] = [team, getPowerScore(team, typeOfPowerscore)]

    #output.write(compName + " PowerScore\n")
    
    teamList.sort(key=lambda x: x[1], reverse=True) # sorts the list based on powerScore from largest to smallest
    teamLibrary = {}
    for team in teamList:
        if team[0][0] != "#":
            #output.write(f"{team[0]} \t {str(team[1])} \n") # writes output to a .txt file
            teamLibrary[team[0]] = team[1]
    #print(teamLibrary)
    return teamLibrary, teamList
    #print("Output has been saved to output.txt")
    #io.showOutput("output.txt")

def runComp(sku, div): # master function for computing a competition powerscore
    name, compInfos, divs, season = apiHandler.getCompInfoBySKU(sku, div) # an api module function that gets the competition info for a comp and division
    print(divs)
    scales = {
        "164" : 0.97,
        "173" : 0.94,
        "181" : 1,
        "190" : 3.2
    }
    for i in range(len(compInfos)):
        fullPSLib, fullPSList = runPowerScore(None, None, div=None, typeOfPowerscore="general", compInfo=compInfos[i], onlyForComp=True, scalingFactor=scales[season])
        fullOPSLib, fullOPSList = runPowerScore(None, None, div=None, typeOfPowerscore="offensive", compInfo=compInfos[i], onlyForComp=True, scalingFactor=scales[season])
        fullDPSLib, fullDPSList = runPowerScore(None, None, div=None, typeOfPowerscore="defensive", compInfo=compInfos[i], onlyForComp=True, scalingFactor=scales[season])

    wb = Workbook() # initializes a new spreadsheet to write the values in for downloading
    ws = wb.active

    newOPSList = []
    newDPSList = []
    row = 1
    ws['A1'] = "Team Number"
    ws['B1'] = "Powerscore"
    ws['C1'] = "Offensive PS"
    ws['D1'] = "Defensive PS"
    for team in fullPSList:
        row += 1
        newOPSList.append([team[0], fullOPSLib[team[0]]]) # reorders the ops and dps lists to be in the same order as the ps list
        newDPSList.append([team[0], fullDPSLib[team[0]]])
        ws[f'A{row}'] = team[0]
        ws[f'B{row}'] = team[1]
        ws[f'C{row}'] = newOPSList[row - 2][1]
        ws[f'D{row}'] = newDPSList[row - 2][1]
    excelFile = BytesIO()
    wb.save(excelFile)
    excelFile.seek(0)

    return [name, fullPSList, newOPSList, newDPSList, divs], excelFile

def runAlgorithm(team, season):
    global accolades, teamname
    scales = {
        "164" : 0.97,
        "173" : 0.94,
        "181" : 1,
        "190" : 3.2
    }

    apiHandler.setDefaultSeason(season)
    accolades = []
    teamname = team

    comps = apiHandler.getCompList(team)

    psList = []
    opsList = []
    dpsList = []

    dashboard = []
    compiledList = asyncApi.getCompiledDataList(team, comps, season)
    comps = comps['data']
    startDate = comps[0]["start"][:10]

    print(compiledList[0][0])
    for comp in range(len(compiledList)):
        division = compiledList[comp]

        date = "Error Fetching Date"
        sku = "Error Fetching SKU"
        for competition in range(len(comps)):
            if comps[competition]["sku"] == compiledList[comp][0]['event']['code']:
                date = comps[competition]["start"][:10]

        # the new powerscore algorithm works for offensive and defensive powerscore. thus the alg is split in three pieces. overall ps is the most important still.
        fullPSLib, fullPSList = runPowerScore(None, None, div=None, typeOfPowerscore="general", compInfo=division, scalingFactor=scales[season])
        newPSLib, newPSList = runPowerScore(None, None, div=None, typeOfPowerscore="general", compInfo=division, onlyForComp=True, scalingFactor=scales[season])
        fullOPSLib, fullOPSList = runPowerScore(None, None, div=None, typeOfPowerscore="offensive", compInfo=division, scalingFactor=scales[season])
        fullDPSLib, fullDPSList = runPowerScore(None, None, div=None, typeOfPowerscore="defensive", compInfo=division, scalingFactor=scales[season])
    
        compPS = fullPSLib[team]
        newPS = newPSLib[team]
        compOPS = fullOPSLib[team]
        compDPS = fullDPSLib[team]

        ### ACCOLADES ###
        if round(newPS) >= 90 and not "First Class" in accolades:
            accolades.append("First Class")
        if round(newPS) >= 90 and ("Signature Event" in compiledList[comp][0]['event']['name']) and not "World Class" in accolades:
            accolades.append("World Class")
        if newPSList.index([team, newPS]) == 0 and not "Top Fragger" in accolades:
            accolades.append("Top Fragger")

        #print(compiledList[comp][0])
        compWeight = 1.0 + 1 * ("Signature Event" in compiledList[comp][0]['event']['name'])
        psList.append([compPS, compWeight, date])
        opsList.append([compOPS, compWeight, getDays(date)])
        dpsList.append([compDPS, compWeight, getDays(date)])

        thisComp = {}
        thisComp['name'] = compiledList[comp][0]['event']['name']
        print(thisComp['name'])
        thisComp['date'] = date
        thisComp['sku'] = compiledList[comp][0]['event']['code']
        thisComp['score'] = round(newPS)

        dashboard.append(thisComp)
    
    ### GENERAL POWERSCORE ###
    summation = 0
    index = 1
    progression = []
    oldTemporalPS = 0
    temporalPS = 0
    for x in psList:
        summation += x[0] * x[1]
        oldTemporalPS = temporalPS
        temporalPS = summation / index
        progression.append([x[2], round((((2/(1 + exp(-0.045 * (temporalPS)))) - 1) / 2 + 0.5) * 1000) / 10])
        index += 1
    plotBytes = createPlot(progression, teamname, startDate)
    plotBytes.seek(0)
    if progression[0][1] < 51 and progression[len(progression) - 1][1] > 70:
        accolades.append("Glow Up")
    sortedProgression = sorted(progression, key=lambda x: x[1], reverse=True) # sorts the list based on powerScore from largest to smallest
    if sortedProgression[0][1] > 80 and progression[len(progression) - 1][1] < 71:
        accolades.append("Fall from Grace")
    careerPS = round((((2/(1 + exp(-0.045 * temporalPS))) - 1) / 2 + 0.5) * 10000) / 100
    #print(careerPS)
    oldCareerPS = round((((2/(1 + exp(-0.045 * oldTemporalPS))) - 1) / 2 + 0.5) * 100)

    nextRank = 0
    prevRank = 0
    for rank in reversed(ranking): # finds the previous and next rank of the player to calculate the xp to next rank
        if careerPS < rank[1]:
            nextRank = rank[1]
    for rank in ranking:
        if careerPS > rank[1]:
            prevRank = rank[1]
    
    xpToNext = round(((nextRank - careerPS) * careerPS) * 0.05) * 100

    percentageXP = round(((careerPS - prevRank) / (nextRank - prevRank)) * 100) / 100

    image = Image.new("RGBA", (100, 10)) # generates progress bar for xp
    for y in range(10):
        for x in range(100):
            if x / 100 < percentageXP:
                image.putpixel((x, y), (132, 238, 255, 255))
            else:
                image.putpixel((x, y), (27, 119, 138, 255))
    imageBytes = BytesIO()
    image.save(imageBytes, "png")
    imageBytes.seek(0)
    rank = giveRanking(careerPS)
    careerPS = round(careerPS)

    ### OFFENSIVE POWERSCORE ###
    summation = 0
    index = 1
    progression = []
    temporalOPS = 0
    for x in opsList:
        summation += x[0] * x[1]
        temporalOPS = summation / index
        progression.append([x[2], round((((2/(1 + exp(-0.045 * (temporalOPS)))) - 1) / 2 + 0.5) * 1000) / 10])
        index += 1
    careerOPS = round((((2/(1 + exp(-0.045 * temporalOPS))) - 1) / 2 + 0.5) * 100)

    ### DEFENSIVE POWERSCORE ###
    summation = 0
    index = 1
    progression = []
    temporalDPS = 0
    for x in dpsList:
        summation += x[0] * x[1]
        temporalDPS = summation / index
        progression.append([x[2], round((((2/(1 + exp(-0.045 * (temporalDPS)))) - 1) / 2 + 0.5) * 1000) / 10])
        index += 1
    careerDPS = round((((2/(1 + exp(-0.045 * temporalDPS))) - 1) / 2 + 0.5) * 100)

    if careerOPS - careerDPS > 5:
        title = "Aggressor"
    elif careerOPS - careerDPS < -5:
        title = "Sentinel"
    else:
        title = "All-Rounder"

    #if "World Class" in accolades:
        #accolades.remove("First Class")
    possibleAccolades = [
        ["World Class", "Attain a Powerscore above 90 at a Signature Event"],
        ["Glow Up", "Climb from below Gold to above Gold in a season"],
        ["Top Fragger", "Attain the highest Powerscore at a competition"],
        ["First Class", "Attain a Powerscore above 90 at a competition"],
        ["Fall from Grace", "Fall from above Diamond to below Diamond in a season"]
    ]
    num = 0
    accolade1 = ["No Accolade", "Earn accolades by playing in tournaments!"]
    accolade2 = ["No Accolade", "Earn accolades by playing in tournaments!"]
    for accolade in possibleAccolades:
        if accolade[0] in accolades:
            if num == 0:
                accolade1 = accolade
            if num == 1:
                accolade2 = accolade
            num += 1

    rankCopy = rank.replace(" ", "")
    badgeFileDir = "newbadges/" + rankCopy + ".png"
    
    # Read the image file
    image_data = open(badgeFileDir, 'rb').read()
    # Encode as base64
    data_uri = base64.b64encode(image_data).decode('utf-8')
    # Create an <img> tag with the base64-encoded image
    badge_tag = data_uri
    
    image_data = plotBytes.getvalue()
    # Encode as base64
    data_uri = base64.b64encode(image_data).decode('utf-8')
    # Create an <img> tag with the base64-encoded image
    graph_tag = data_uri

    image_data = imageBytes.getvalue()
    # Encode as base64
    data_uri = base64.b64encode(image_data).decode('utf-8')
    # Create an <img> tag with the base64-encoded image
    bar_tag = data_uri
    
    return [team, careerPS, oldCareerPS, rank, careerOPS, careerDPS, title, accolade1, accolade2, badge_tag, graph_tag, xpToNext, bar_tag, dashboard]

    #print(comps["data"][comp]["id"])